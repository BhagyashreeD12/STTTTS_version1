"""
flow_parser.py — Parse VoiceChat_QuestionSequenceAndProb.xlsx
              → data/insurance_flow.json

Uses the "Voice Agent Questions" sheet (the voice-specific version of the flow).

Run once (or whenever the Excel is updated):
    .venv\\Scripts\\python.exe scripts/flow_parser.py
"""

import json
import re
from pathlib import Path

import openpyxl

# ── Paths ─────────────────────────────────────────────────────────────────────
_HERE = Path(__file__).parent
XLSX_PATH = _HERE.parent.parent / "data" / "VoiceChat_QuestionSequenceAndProb.xlsx"
OUTPUT_JSON = _HERE.parent.parent / "data" / "insurance_flow.json"
SHEET_NAME = "Voice Agent Questions"

# ── Steps that belong to the driver loop (repeat for each driver) ─────────────
DRIVER_LOOP_STEPS = {"2.2", "2.3", "2.4", "2.4a", "2.5", "2.6", "2.7"}

# ── Steps that belong to the vehicle loop (repeat for each vehicle) ───────────
VEHICLE_LOOP_STEPS = {"3.3", "3.4", "3.5", "3.6", "3.7", "3.8", "3.9", "3.10"}

# ── Known branching logic names keyed by step ID ──────────────────────────────
BRANCH_LOGIC: dict[str, str] = {
    "1.1": "consent_check",
    "1.4": "modifications_check",
    "2.1": "driver_loop_start",
    "2.2": "registered_owner_check",
    "2.4": "applicant_check",
    "3.1": "vehicle_loop_start",
    "3.3": "principal_driver_check",
    "3.4": "ownership_type_check",
    "7.1": "confirmation_check",
    "7.2": "farewell",
}


def _parse_options(raw) -> list[str]:
    if not raw or str(raw).strip() in ("-", ""):
        return []
    parts = re.split(r"\s*/\s*|,\s*", str(raw).strip())
    return [p.strip() for p in parts if p.strip()]


def parse_excel() -> dict:
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_idx = {h: i + 1 for i, h in enumerate(headers)}

    def cell(row: int, header: str):
        return ws.cell(row, col_idx[header]).value

    steps: list[dict] = []
    current_block: str | None = None

    for r in range(2, ws.max_row + 1):
        q_no = cell(r, "Question No.")
        block = cell(r, "Blocks")
        q_text = cell(r, "Questions")
        opts = cell(r, "Options")
        exp = cell(r, "Expected Answer")
        cond = cell(r, "Conditions")

        # Skip fully empty rows
        if all(v is None for v in [q_no, block, q_text, opts, exp, cond]):
            continue

        # Track current block (header rows)
        if block:
            current_block = str(block).strip()

        # Row 10 has no Question No. — it is the "relationship" sub-question under 2.4
        if q_no is None and q_text:
            q_no = "2.4a"

        if not q_text:
            continue

        sid = str(q_no).strip()
        question = str(q_text).strip().replace("\n", " ")

        step: dict = {
            "id": sid,
            "block": current_block or "UNKNOWN",
            "question": question,
            "options": _parse_options(opts),
            "expected": str(exp).strip() if exp else None,
            "conditions": str(cond).strip() if cond else None,
            "branch_logic": BRANCH_LOGIC.get(sid),
            "in_driver_loop": sid in DRIVER_LOOP_STEPS,
            "in_vehicle_loop": sid in VEHICLE_LOOP_STEPS,
        }
        steps.append(step)

    return {
        "version": "1.0",
        "source_file": XLSX_PATH.name,
        "source_sheet": SHEET_NAME,
        "steps": steps,
    }


def main() -> None:
    print(f"Reading: {XLSX_PATH}")
    flow = parse_excel()

    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(
        json.dumps(flow, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    print(f"Saved:   {OUTPUT_JSON}")
    print(f"Steps:   {len(flow['steps'])}\n")
    print("Step summary:")

    current_block = None
    for s in flow["steps"]:
        if s["block"] != current_block:
            current_block = s["block"]
            print(f"\n  [{current_block}]")
        tags = []
        if s["in_driver_loop"]:
            tags.append("driver-loop")
        if s["in_vehicle_loop"]:
            tags.append("vehicle-loop")
        if s["branch_logic"]:
            tags.append(f"branch:{s['branch_logic']}")
        tag_str = "  " + "  ".join(tags) if tags else ""
        print(f"    {s['id']:6s}  {s['question'][:65]!r}{tag_str}")


if __name__ == "__main__":
    main()
